from numpy import save
import openpyxl
from openpyxl import Workbook
total_files = 37
# new excel
wb_dest = Workbook()
ws_dest = wb_dest.active

dest_row = 1
dest_col = 1
# existing file
for excel_count in range(0, total_files):
    excel_file = f"reportTable ({excel_count})"
    # print(excel_file) #debug
    wb = openpyxl.load_workbook(f"excel\\{excel_file}.xlsx")
    ws = wb["reportTable"]
    row = -1
    for excel_row in ws.iter_rows(min_row=6, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
        for excel_cell in excel_row:
            if excel_cell.value == "TOYOTA KIRLOSKAR MOTOR PVT LTD":
                row = 5
                # print (excel_cell.value, end="\n")
                continue
            if row <= 5 and row >= 0:
                if dest_col == 6:
                    dest_col = 1
                    dest_row += 1
                ws_dest.cell(dest_row, dest_col, value=int(excel_cell.value))
                dest_col += 1
                # print (excel_cell.value, end=" ")
                row -= 1
        if row == 0:
                break

ws_dest["B50"] = "= SUM(B1:B49)"
ws_dest["A50"] = "LMV = "
wb_dest.save("Final.xlsx")  
            
